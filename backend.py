from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
import requests
import json
from typing import List, Dict, Any, Optional, Tuple
from pydantic import BaseModel
import warnings
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime, timedelta
import tempfile
import os
import sys
import webbrowser
import threading
import hashlib

# HTTPS 인증서 경고 무시
warnings.simplefilter("ignore", InsecureRequestWarning)

app = FastAPI(title="Nutanix Cluster Manager API")

# ============ 캐싱 시스템 ============
# 캐시 저장소: {cache_key: (data, timestamp)}
cache_storage: Dict[str, Tuple[List[Dict[str, Any]], datetime]] = {}

# 캐시 TTL (Time To Live) - 5분
CACHE_TTL_MINUTES = 5

def generate_cache_key(cluster_id: str, category: str, **params) -> str:
    """캐시 키 생성 - 클러스터 ID, 카테고리, 추가 파라미터 기반"""
    # Performance의 경우 시간 파라미터도 포함
    if category == "Performance":
        key_parts = [
            cluster_id,
            category,
            str(params.get('startTime', '')),
            str(params.get('endTime', '')),
            str(params.get('interval', '')),
            str(params.get('aggregationType', ''))
        ]
    elif category == "Resources":
        key_parts = [
            cluster_id,
            category,
            str(params.get('ratio', '')),
            str(params.get('rf', ''))
        ]
    else:
        key_parts = [cluster_id, category]
    
    key_string = "|".join(key_parts)
    return hashlib.md5(key_string.encode()).hexdigest()

def get_cached_data(cache_key: str) -> Optional[List[Dict[str, Any]]]:
    """캐시에서 데이터 가져오기 - TTL 체크"""
    if cache_key in cache_storage:
        data, timestamp = cache_storage[cache_key]
        # TTL 체크
        if datetime.now() - timestamp < timedelta(minutes=CACHE_TTL_MINUTES):
            print(f"[CACHE HIT] Key: {cache_key}")
            return data
        else:
            # 만료된 캐시 삭제
            print(f"[CACHE EXPIRED] Key: {cache_key}")
            del cache_storage[cache_key]
    
    print(f"[CACHE MISS] Key: {cache_key}")
    return None

def set_cached_data(cache_key: str, data: List[Dict[str, Any]]) -> None:
    """캐시에 데이터 저장"""
    cache_storage[cache_key] = (data, datetime.now())
    print(f"[CACHE SET] Key: {cache_key}, Items: {len(data)}")

def clear_cache(cluster_id: Optional[str] = None) -> None:
    """캐시 삭제 - 특정 클러스터 또는 전체"""
    if cluster_id:
        # 특정 클러스터의 캐시만 삭제
        keys_to_delete = [k for k in cache_storage.keys() if k.startswith(cluster_id)]
        for key in keys_to_delete:
            del cache_storage[key]
        print(f"[CACHE CLEAR] Cluster: {cluster_id}")
    else:
        # 전체 캐시 삭제
        cache_storage.clear()
        print(f"[CACHE CLEAR] All cache cleared")
# ============ 캐싱 시스템 끝 ============

app = FastAPI(title="Nutanix Cluster Manager API")

# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # 개발용 - 운영에서는 특정 도메인만 허용
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 정적 파일 경로 설정 (PyInstaller 고려)
if getattr(sys, 'frozen', False):
    # PyInstaller로 실행 중
    BASE_DIR = sys._MEIPASS
else:
    # 일반 Python으로 실행 중
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

STATIC_DIR = os.path.join(BASE_DIR, "dist")

# 정적 파일 서빙
if os.path.exists(STATIC_DIR):
    # assets 폴더
    app.mount("/assets", StaticFiles(directory=os.path.join(STATIC_DIR, "assets")), name="assets")
    # 루트 레벨의 정적 파일들 (이미지 등)
    app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# 데이터 모델
class ClusterConfig(BaseModel):
    id: str
    name: str
    ip: str
    username: str
    password: str
    isVerified: bool = False
    type: str = "PE"
    apiVersion: str = "v2.0"

class DataRow(BaseModel):
    clusterName: str = ""
    name: str = ""
    uuid: str = ""
    powerState: str = ""
    ipAddresses: str = ""
    numVcpus: int = 0
    memoryMb: int = 0

# Nutanix API 헬퍼 함수들
def get_api_url(ip: str, endpoint: str) -> str:
    return f"https://{ip}:9440/api/nutanix/v2.0{endpoint}"

def nutanix_fetch(url: str, cluster: ClusterConfig, timeout_secs: int = 15) -> Dict[str, Any]:
    auth = (cluster.username, cluster.password)

    try:
        response = requests.get(
            url,
            auth=auth,
            verify=False,
            timeout=timeout_secs
        )

        if not response.ok:
            raise HTTPException(
                status_code=response.status_code,
                detail=f"Nutanix API Error: {response.status_code} {response.reason}"
            )

        return response.json()
    except requests.exceptions.Timeout:
        raise HTTPException(status_code=408, detail=f"Timeout: Server at {cluster.ip} did not respond in {timeout_secs}s")
    except requests.exceptions.ConnectionError:
        raise HTTPException(status_code=503, detail=f"Connection failed: Cannot connect to {cluster.ip}")

def get_cluster_name(cluster: ClusterConfig) -> str:
    """클러스터 이름 가져오기"""
    url = get_api_url(cluster.ip, '/cluster')
    try:
        json_data = nutanix_fetch(url, cluster)
        return json_data.get("name", "Unknown")
    except Exception:
        return "Unknown"

# API 엔드포인트들
@app.post("/api/verify-cluster")
async def verify_cluster(cluster: ClusterConfig) -> dict:
    """클러스터 연결 검증 및 실제 클러스터 이름 반환"""
    url = get_api_url(cluster.ip, '/clusters')
    try:
        nutanix_fetch(url, cluster, timeout_secs=10)
        # 실제 클러스터 이름 가져오기
        actual_cluster_name = get_cluster_name(cluster)
        return {"success": True, "clusterName": actual_cluster_name}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.post("/api/clear-cache")
async def clear_cache_endpoint(cluster_id: Optional[str] = None) -> dict:
    """캐시 삭제 API - 특정 클러스터 또는 전체"""
    clear_cache(cluster_id)
    return {"status": "success", "message": f"Cache cleared for cluster: {cluster_id}" if cluster_id else "All cache cleared"}

@app.get("/api/cache-stats")
async def cache_stats() -> dict:
    """캐시 통계 조회"""
    total_items = len(cache_storage)
    cache_info = []
    
    for key, (data, timestamp) in cache_storage.items():
        age_seconds = (datetime.now() - timestamp).total_seconds()
        cache_info.append({
            "key": key[:16] + "...",  # 키 일부만 표시
            "items": len(data),
            "age_seconds": int(age_seconds),
            "expires_in": int(CACHE_TTL_MINUTES * 60 - age_seconds)
        })
    
    return {
        "total_cached_keys": total_items,
        "ttl_minutes": CACHE_TTL_MINUTES,
        "cache_details": cache_info
    }

@app.post("/api/fetch-data")
async def fetch_data(
    cluster: ClusterConfig, 
    category: str = "VM",
    startTime: Optional[str] = None,
    endTime: Optional[str] = None,
    interval: Optional[int] = 30,
    aggregationType: Optional[str] = "average",
    ratio: Optional[int] = 3,
    rf: Optional[int] = 2
) -> List[Dict[str, Any]]:
    """데이터 조회 (캐싱 적용)"""
    
    # 캐시 키 생성
    cache_key = generate_cache_key(
        cluster.id, 
        category,
        startTime=startTime,
        endTime=endTime,
        interval=interval,
        aggregationType=aggregationType,
        ratio=ratio,
        rf=rf
    )
    
    # 캐시 확인
    cached_data = get_cached_data(cache_key)
    if cached_data is not None:
        return cached_data
    
    # 캐시 미스 - API 호출하여 데이터 가져오기
    result = []
    
    if category == "VM":
        # 클러스터 이름 가져오기
        cluster_name = get_cluster_name(cluster)

        # VM 목록 가져오기 - include_vm_nic_config=true, include_vm_disk_config=true로 NIC 및 디스크 상세정보 포함
        url = get_api_url(cluster.ip, '/vms?include_cvm=true&include_vm_nic_config=true&include_vm_disk_config=true')
        json_data = nutanix_fetch(url, cluster)
        entities = json_data.get("entities", [])

        result = []
        for vm in entities:
            vm_uuid = vm.get("uuid", "")
            vm_name = vm.get("vmName", "") or vm.get("name", "")
            vm_power_state = vm.get("power_state", "UNKNOWN")
            vm_nics = vm.get("vm_nics", [])
            
            # NIC 정보 처리 - MAC과 IP를 분리해서 표시
            mac_addresses = []
            ip_addresses = []
            for nic in vm_nics:
                mac = nic.get("mac_address", "")
                ip = nic.get("ip_address", "")
                
                if mac:
                    mac_addresses.append(mac)
                
                # IP는 있으면 IP, 없으면 'none-ip-setting'
                ip_display = ip if ip else "none-ip-setting"
                if mac:  # MAC이 있는 경우에만 IP 추가
                    ip_addresses.append(ip_display)
            
            # 개행으로 구분하여 표시
            mac_str = "\n".join(mac_addresses) if mac_addresses else ""
            ip_str = "\n".join(ip_addresses) if ip_addresses else ""
            
            # vDisk 정보 처리
            vm_disks = vm.get("vm_disk_info", [])
            disk_list = []
            for disk in vm_disks:
                disk_label = disk.get("disk_address", {}).get("disk_label", "N/A")
                size_bytes = disk.get("size", 0)
                # Byte를 GiB로 변환
                size_gib = size_bytes / (1024**3) if size_bytes else 0
                disk_str = f"{disk_label} : {size_gib:.0f} GiB"
                disk_list.append(disk_str)
            
            vdisk_str = "\n".join(disk_list) if disk_list else ""
            
            num_vcpus = vm.get("num_vcpus", 0)
            memory_mb = vm.get("memory_mb", 0)

            result.append({
                "clusterName": cluster_name,
                "name": vm_name,
                "uuid": vm_uuid,
                "powerState": vm_power_state,
                "macAddress": mac_str,
                "ipAddresses": ip_str,
                "vDisk": vdisk_str,
                "numVcpus": num_vcpus,
                "memoryMb": memory_mb,
            })

        # 캐시 저장
        set_cached_data(cache_key, result)
        return result

    elif category == "Performance":
        # Performance 데이터 가져오기
        if not startTime or not endTime:
            raise HTTPException(status_code=400, detail="startTime and endTime are required for Performance category")
        
        # datetime 문자열을 microseconds로 변환
        try:
            start_dt = datetime.fromisoformat(startTime.replace('Z', '+00:00'))
            end_dt = datetime.fromisoformat(endTime.replace('Z', '+00:00'))
            start_usec = int(start_dt.timestamp() * 1000000)
            end_usec = int(end_dt.timestamp() * 1000000)
        except ValueError as e:
            raise HTTPException(status_code=400, detail=f"Invalid datetime format: {e}")
        
        # 클러스터 정보 가져오기
        cluster_url = get_api_url(cluster.ip, '/clusters')
        cluster_data = nutanix_fetch(cluster_url, cluster)
        cluster_entity = cluster_data.get("entities", [{}])[0]
        cluster_id = cluster_entity.get("uuid", "")
        cluster_name = cluster_entity.get("name", "Unknown")
        
        # 호스트 목록 가져오기
        host_url = get_api_url(cluster.ip, '/hosts')
        host_data = nutanix_fetch(host_url, cluster)
        host_entities = host_data.get("entities", [])
        
        result = []
        
        # v1 API 사용 (첨부 파일 참고)
        base_stats_url = f"https://{cluster.ip}:9440/PrismGateway/services/rest/v1"
        
        # 메트릭 정의
        metrics = [
            "controller_num_iops",
            "controller_avg_io_latency_usecs", 
            "controller_io_bandwidth_kBps",
            "hypervisor_cpu_usage_ppm",
            "hypervisor_memory_usage_ppm"
        ]
        
        # 클러스터 통계 가져오기 - 모든 메트릭을 한 번에 요청
        cluster_stats = {}
        try:
            # 모든 메트릭을 쉼표로 연결하여 한 번에 요청
            metrics_str = ','.join(metrics)
            params = {
                'metrics': metrics_str,
                'startTimeInUsecs': start_usec,
                'endTimeInUsecs': end_usec,
                'intervalInSecs': interval
            }
            
            stats_url = f"{base_stats_url}/clusters/{cluster_id}/stats/"
            response = requests.get(
                stats_url,
                params=params,
                auth=(cluster.username, cluster.password),
                verify=False,
                timeout=30
            )
            
            if response.ok:
                stats_data = response.json()
                stats_responses = stats_data.get("statsSpecificResponses", [])
                
                # 각 메트릭별로 응답 처리
                for i, metric in enumerate(metrics):
                    if i < len(stats_responses):
                        values = stats_responses[i].get("values", [])
                        
                        if values:
                            if aggregationType == "max":
                                cluster_stats[metric] = max(values)
                            elif aggregationType == "min":
                                cluster_stats[metric] = min(values)
                            else:  # average
                                cluster_stats[metric] = sum(values) / len(values)
                        else:
                            cluster_stats[metric] = 0
                    else:
                        cluster_stats[metric] = 0
            else:
                # 실패 시 모든 메트릭을 0으로 설정
                for metric in metrics:
                    cluster_stats[metric] = 0
        except Exception as e:
            print(f"Error fetching cluster stats: {e}")
            # 에러 시 모든 메트릭을 0으로 설정
            for metric in metrics:
                cluster_stats[metric] = 0
        
        # 클러스터 행 추가
        cluster_row = {
            "entityType": "cluster",
            "entityName": cluster_name,
            "iops": f"{int(cluster_stats.get('controller_num_iops', 0)):,}",
            "latency": f"{cluster_stats.get('controller_avg_io_latency_usecs', 0) / 1000:.2f}",  # usec -> ms
            "bandwidth": f"{cluster_stats.get('controller_io_bandwidth_kBps', 0) / 1024:.2f} MB/s",  # KB/s -> MB/s
            "cpuUsage": f"{cluster_stats.get('hypervisor_cpu_usage_ppm', 0) / 10000:.2f}",  # ppm -> %
            "memoryUsage": f"{cluster_stats.get('hypervisor_memory_usage_ppm', 0) / 10000:.2f}",  # ppm -> %
        }
        result.append(cluster_row)
        
        # 각 호스트의 통계 가져오기 - 모든 메트릭을 한 번에 요청
        for host in host_entities:
            host_id = host.get("uuid", "")
            host_name = host.get("name", "Unknown")
            
            host_stats = {}
            try:
                # 모든 메트릭을 쉼표로 연결하여 한 번에 요청
                metrics_str = ','.join(metrics)
                params = {
                    'metrics': metrics_str,
                    'startTimeInUsecs': start_usec,
                    'endTimeInUsecs': end_usec,
                    'intervalInSecs': interval
                }
                
                stats_url = f"{base_stats_url}/hosts/{host_id}/stats/"
                response = requests.get(
                    stats_url,
                    params=params,
                    auth=(cluster.username, cluster.password),
                    verify=False,
                    timeout=30
                )
                
                if response.ok:
                    stats_data = response.json()
                    stats_responses = stats_data.get("statsSpecificResponses", [])
                    
                    # 각 메트릭별로 응답 처리
                    for i, metric in enumerate(metrics):
                        if i < len(stats_responses):
                            values = stats_responses[i].get("values", [])
                            
                            if values:
                                if aggregationType == "max":
                                    host_stats[metric] = max(values)
                                elif aggregationType == "min":
                                    host_stats[metric] = min(values)
                                else:  # average
                                    host_stats[metric] = sum(values) / len(values)
                            else:
                                host_stats[metric] = 0
                        else:
                            host_stats[metric] = 0
                else:
                    # 실패 시 모든 메트릭을 0으로 설정
                    for metric in metrics:
                        host_stats[metric] = 0
            except Exception as e:
                print(f"Error fetching stats for host {host_name}: {e}")
                # 에러 시 모든 메트릭을 0으로 설정
                for metric in metrics:
                    host_stats[metric] = 0
            
            # 호스트 행 추가
            host_row = {
                "entityType": "host",
                "entityName": host_name,
                "parentCluster": cluster_name,
                "iops": f"{int(host_stats.get('controller_num_iops', 0)):,}",
                "latency": f"{host_stats.get('controller_avg_io_latency_usecs', 0) / 1000:.2f}",
                "bandwidth": f"{host_stats.get('controller_io_bandwidth_kBps', 0) / 1024:.2f} MB/s",
                "cpuUsage": f"{host_stats.get('hypervisor_cpu_usage_ppm', 0) / 10000:.2f}",
                "memoryUsage": f"{host_stats.get('hypervisor_memory_usage_ppm', 0) / 10000:.2f}",
            }
            result.append(host_row)
        
        # 캐시 저장
        set_cached_data(cache_key, result)
        return result

    # 다른 카테고리는 간단히 구현
    elif category == "Hardware":
        # 호스트 정보 가져오기
        url = get_api_url(cluster.ip, '/hosts')
        json_data = nutanix_fetch(url, cluster)
        entities = json_data.get("entities", [])
        
        # 노드 UUID별 디스크 정보 맵 생성 (개별 디스크 목록)
        disk_info_map = {}
        disk_model_map = {}
        try:
            # 디스크 정보 가져오기
            disk_url = get_api_url(cluster.ip, '/disks')
            disk_data = nutanix_fetch(disk_url, cluster)
            disk_entities = disk_data.get("entities", [])
            
            # 노드 UUID별로 디스크 정보 수집 (각 디스크를 개별적으로 저장)
            for disk in disk_entities:
                node_uuid = disk.get("node_uuid", "")
                if node_uuid:
                    if node_uuid not in disk_info_map:
                        disk_info_map[node_uuid] = []
                        disk_model_map[node_uuid] = []
                    
                    # 디스크 타입과 용량
                    disk_type = disk.get("storage_tier_name", "UNKNOWN")
                    capacity_bytes = disk.get("disk_size", 0)
                    
                    # 용량을 TB, GB, MB로 변환
                    if capacity_bytes >= 1024**4:
                        capacity_str = f"{capacity_bytes / (1024**4):.2f}TB".rstrip('0').rstrip('.')
                    elif capacity_bytes >= 1024**3:
                        capacity_str = f"{capacity_bytes / (1024**3):.2f}GB".rstrip('0').rstrip('.')
                    else:
                        capacity_str = f"{capacity_bytes / (1024**2):.2f}MB".rstrip('0').rstrip('.')
                    
                    disk_str = f"{capacity_str} {disk_type}"
                    disk_info_map[node_uuid].append(disk_str)
                    
                    # 디스크 모델 정보
                    disk_model = disk.get("disk_hardware_config", {}).get("model", "N/A")
                    disk_model_map[node_uuid].append(disk_model)
        except Exception:
            # 디스크 정보를 가져올 수 없으면 무시
            pass

        result = []
        for host in entities:
            host_uuid = host.get("uuid", "")
            
            # 디스크 정보 포맷팅 (각 디스크를 개별 라인으로)
            disk_str = ""
            disk_model_str = ""
            if host_uuid in disk_info_map:
                disk_str = "\n".join(disk_info_map[host_uuid])
                disk_model_str = "\n".join(disk_model_map[host_uuid])
            
            result.append({
                "hostName": host.get("name", "Unknown"),
                "serial": host.get("serial", "N/A"),
                "model": host.get("block_model_name", host.get("model", "N/A")),
                "cpuModel": host.get("cpu_model", "N/A"),
                "numCores": host.get("num_cpu_cores", 0),
                "memoryCapacity": int(host.get("memory_capacity_in_bytes", 0) / (1024**3)) if host.get("memory_capacity_in_bytes") else 0,
                "blockSerial": host.get("block_serial", "N/A"),
                "disk": disk_str,
                "diskModel": disk_model_str,
            })
        
        # 캐시 저장
        set_cached_data(cache_key, result)
        return result
    
    elif category == "Resources":
        # Resources 카테고리 - CPU 및 Memory 분석
        
        # 호스트 정보 가져오기
        host_url = get_api_url(cluster.ip, '/hosts')
        host_data = nutanix_fetch(host_url, cluster)
        host_entities = host_data.get("entities", [])
        
        # VM 목록 가져오기
        vm_url = get_api_url(cluster.ip, '/vms?include_cvm=true')
        vm_data = nutanix_fetch(vm_url, cluster)
        vm_entities = vm_data.get("entities", [])
        
        # 호스트별 VM 정보 집계
        host_vm_map = {}
        for vm in vm_entities:
            host_uuid = vm.get("host_uuid", "")
            if not host_uuid:
                continue
            
            if host_uuid not in host_vm_map:
                host_vm_map[host_uuid] = {
                    "vms": [],
                    "powered_on_vms": []
                }
            
            host_vm_map[host_uuid]["vms"].append(vm)
            if vm.get("power_state") == "on":
                host_vm_map[host_uuid]["powered_on_vms"].append(vm)
        
        # CPU 테이블 데이터 생성
        cpu_rows = []
        memory_rows = []
        
        for host in host_entities:
            host_uuid = host.get("uuid", "")
            host_name = host.get("name", "Unknown")
            num_cores = host.get("num_cpu_cores", 0)
            memory_capacity_bytes = host.get("memory_capacity_in_bytes", 0)
            memory_capacity_gib = int(memory_capacity_bytes / (1024**3)) if memory_capacity_bytes else 0
            
            vms = host_vm_map.get(host_uuid, {}).get("vms", [])
            powered_on_vms = host_vm_map.get(host_uuid, {}).get("powered_on_vms", [])
            
            vm_count = len(vms)
            
            # CPU 계산
            numa_over_cpu = 0
            use_vcores = 0
            for vm in powered_on_vms:
                vcpus = vm.get("num_vcpus", 0)
                use_vcores += vcpus
                # NUMA over: vCore > pCore/2
                if vcpus > (num_cores / 2):
                    numa_over_cpu += 1
            
            recommend_vcore_ratio = f"1:{ratio}"
            current_vcore_ratio = f"1:{use_vcores / num_cores:.1f}" if num_cores > 0 else "1:0.0"
            current_vcore_ratio_value = use_vcores / num_cores if num_cores > 0 else 0
            
            cpu_result = "PASS" if current_vcore_ratio_value <= ratio else "FAIL"
            
            cpu_rows.append({
                "table": "CPU",
                "hostName": host_name,
                "vmCount": vm_count,
                "numaOver": numa_over_cpu,
                "pCore": num_cores,
                "useVCore": use_vcores,
                "recommendVcoreRatio": recommend_vcore_ratio,
                "currentVcoreRatio": current_vcore_ratio,
                "result": cpu_result
            })
            
            # Memory 계산
            numa_over_memory = 0
            use_memory_gib = 0
            for vm in powered_on_vms:
                memory_mb = vm.get("memory_mb", 0)
                memory_gib = memory_mb / 1024
                use_memory_gib += memory_gib
                # NUMA over: Memory > 물리서버 Memory/2
                if memory_gib > (memory_capacity_gib / 2):
                    numa_over_memory += 1
            
            memory_rows.append({
                "table": "Memory",
                "hostName": host_name,
                "vmCount": vm_count,
                "numaOver": numa_over_memory,
                "memoryGiB": memory_capacity_gib,
                "useMemoryGiB": int(use_memory_gib),
                "recommendUse": 0,  # Total 계산 후 업데이트
                "availableMemory": 0,  # Total 계산 후 업데이트
                "result": ""  # Total 계산 후 업데이트
            })
        
        # CPU Total 계산
        total_vm_count = sum(row["vmCount"] for row in cpu_rows)
        total_numa_over_cpu = sum(row["numaOver"] for row in cpu_rows)
        total_pcore = sum(row["pCore"] for row in cpu_rows)
        total_use_vcore = sum(row["useVCore"] for row in cpu_rows)
        avg_current_vcore_ratio = total_use_vcore / total_pcore if total_pcore > 0 else 0
        total_cpu_result = "PASS" if avg_current_vcore_ratio <= ratio else "FAIL"
        
        cpu_rows.append({
            "table": "CPU",
            "hostName": "Total",
            "vmCount": total_vm_count,
            "numaOver": total_numa_over_cpu,
            "pCore": total_pcore,
            "useVCore": total_use_vcore,
            "recommendVcoreRatio": f"1:{ratio}",
            "currentVcoreRatio": f"1:{avg_current_vcore_ratio:.2f}",
            "result": total_cpu_result
        })
        
        # Memory Total 계산
        total_memory_gib = sum(row["memoryGiB"] for row in memory_rows)
        total_use_memory_gib = sum(row["useMemoryGiB"] for row in memory_rows)
        total_numa_over_memory = sum(row["numaOver"] for row in memory_rows)
        
        # Recommend Use 계산 (RF에 따라)
        memory_capacities = [row["memoryGiB"] for row in memory_rows]
        memory_capacities_sorted = sorted(memory_capacities, reverse=True)
        
        if rf == 2:
            # RF2: 가장 큰 호스트 1개 제외
            total_recommend_use = total_memory_gib - (memory_capacities_sorted[0] if memory_capacities_sorted else 0)
        else:  # rf == 1 or 3
            # RF3: 가장 큰 호스트 2개 제외
            total_recommend_use = total_memory_gib - sum(memory_capacities_sorted[:2]) if len(memory_capacities_sorted) >= 2 else 0
        
        # 각 호스트의 Recommend Use 계산 (Total 비율 적용)
        recommend_percent = (total_recommend_use / total_memory_gib * 100) if total_memory_gib > 0 else 0
        for row in memory_rows:
            row["recommendUse"] = int(row["memoryGiB"] * recommend_percent / 100)
            row["availableMemory"] = row["recommendUse"] - row["useMemoryGiB"]
            row["result"] = "FAIL" if row["availableMemory"] <= 0 else "PASS"
        
        # Memory Total 행 추가
        available_memory_percent = int(total_recommend_use / total_memory_gib * 100) if total_memory_gib > 0 else 0
        total_memory_result = "PASS" if total_use_memory_gib < total_recommend_use else "FAIL"
        
        memory_rows.append({
            "table": "Memory",
            "hostName": "Total",
            "vmCount": total_vm_count,
            "numaOver": total_numa_over_memory,
            "memoryGiB": total_memory_gib,
            "useMemoryGiB": total_use_memory_gib,
            "recommendUse": total_recommend_use,
            "availableMemory": f"{available_memory_percent}%",
            "result": total_memory_result
        })
        
        # CPU와 Memory 데이터 병합
        result = cpu_rows + memory_rows
        
        # 캐시 저장
        set_cached_data(cache_key, result)
        return result

    else:
        return []

@app.post("/api/export-xlsx")
async def export_xlsx(request_data: dict):
    """엑셀 형식으로 데이터 내보내기"""
    try:
        is_resources = request_data.get("isResources", False)
        filename = request_data.get("filename", "export")

        # 워크북 생성
        wb = Workbook()
        
        # Resources 특별 처리
        if is_resources:
            cpu_data = request_data.get("cpuData", [])
            memory_data = request_data.get("memoryData", [])
            cpu_fields = request_data.get("cpuFields", [])
            memory_fields = request_data.get("memoryFields", [])
            cpu_field_labels = request_data.get("cpuFieldLabels", {})
            memory_field_labels = request_data.get("memoryFieldLabels", {})
            
            # 테두리 및 정렬 스타일
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            data_font = Font(size=10)
            
            # CPU 시트
            ws_cpu = wb.active
            ws_cpu.title = "CPU"
            
            # CPU 헤더
            for col_idx, field in enumerate(cpu_fields, 1):
                cell = ws_cpu.cell(row=1, column=col_idx)
                cell.value = cpu_field_labels.get(field, field)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # CPU 데이터
            for row_idx, row in enumerate(cpu_data, 2):
                for col_idx, field in enumerate(cpu_fields, 1):
                    cell = ws_cpu.cell(row=row_idx, column=col_idx)
                    cell.value = row.get(field, "")
                    cell.number_format = '@'
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.font = data_font
            
            # CPU 열 너비 조정
            for col_idx in range(1, len(cpu_fields) + 1):
                column_letter = chr(64 + col_idx)
                ws_cpu.column_dimensions[column_letter].width = 20
            
            # Memory 시트
            ws_memory = wb.create_sheet(title="Memory")
            
            # Memory 헤더
            for col_idx, field in enumerate(memory_fields, 1):
                cell = ws_memory.cell(row=1, column=col_idx)
                cell.value = memory_field_labels.get(field, field)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border
            
            # Memory 데이터
            for row_idx, row in enumerate(memory_data, 2):
                for col_idx, field in enumerate(memory_fields, 1):
                    cell = ws_memory.cell(row=row_idx, column=col_idx)
                    cell.value = row.get(field, "")
                    cell.number_format = '@'
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.font = data_font
            
            # Memory 열 너비 조정
            for col_idx in range(1, len(memory_fields) + 1):
                column_letter = chr(64 + col_idx)
                ws_memory.column_dimensions[column_letter].width = 20
        
        else:
            # 기존 로직 (VM, Hardware, Performance)
            data = request_data.get("data", [])
            fields = request_data.get("fields", [])
            field_labels = request_data.get("fieldLabels", {})
            
            ws = wb.active
            ws.title = "Data"

            # 테두리 스타일 정의
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 중앙 정렬 (수평 + 수직)
            center_alignment = Alignment(
                horizontal='center',
                vertical='center',
                wrap_text=True
            )

            # 헤더 행 추가
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=10)
            
            for col_idx, field in enumerate(fields, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = field_labels.get(field, field)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
                cell.border = thin_border

            # 데이터 행 추가
            data_font = Font(size=10)
            for row_idx, row in enumerate(data, 2):
                for col_idx, field in enumerate(fields, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    value = row.get(field, "")
                    cell.value = value
                    # 모든 셀을 텍스트 형식으로 설정
                    cell.number_format = '@'
                    cell.alignment = center_alignment
                    cell.border = thin_border
                    cell.font = data_font

            # 열 너비 자동 조정
            for col_idx, field in enumerate(fields, 1):
                column_letter = chr(64 + col_idx)  # A=65, B=66, etc.
                ws.column_dimensions[column_letter].width = 20

            # 첫 행에 필터 설정
            if fields:
                last_col = chr(64 + len(fields))
                ws.auto_filter.ref = f"A1:{last_col}1"

        # 임시 파일로 저장
        temp_file = tempfile.NamedTemporaryFile(mode='w+b', delete=False, suffix='.xlsx')
        temp_path = temp_file.name
        temp_file.close()

        wb.save(temp_path)

        # 파일 응답
        return FileResponse(
            path=temp_path,
            filename=f"{filename}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/")
async def root():
    """루트 경로 - index.html 반환"""
    index_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    return {"message": "Nutanix Cluster Manager API", "status": "running"}

@app.get("/{full_path:path}")
async def serve_spa(full_path: str):
    """SPA용 fallback - 모든 경로를 index.html로"""
    # API 경로는 제외
    if full_path.startswith("api/"):
        raise HTTPException(status_code=404, detail="Not found")
    
    # 정적 파일 직접 서빙 (.jpg, .png, .css, .js 등)
    if "." in full_path.split("/")[-1]:
        file_path = os.path.join(STATIC_DIR, full_path)
        if os.path.exists(file_path):
            return FileResponse(file_path)
    
    # 나머지는 SPA fallback
    index_path = os.path.join(STATIC_DIR, "index.html")
    if os.path.exists(index_path):
        return FileResponse(index_path)
    raise HTTPException(status_code=404, detail="Not found")

def open_browser():
    """브라우저 자동 열기"""
    webbrowser.open('http://localhost:8000')

if __name__ == "__main__":
    import uvicorn
    # 브라우저 자동 열기 (1초 후)
    timer = threading.Timer(1.5, open_browser)
    timer.daemon = True
    timer.start()
    
    uvicorn.run(app, host="0.0.0.0", port=8000)
